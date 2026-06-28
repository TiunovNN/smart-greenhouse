#!/usr/bin/env python3
"""Generate smart-greenhouse BOM Excel/CSV from design doc data."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_XLSX = REPO_ROOT / "docs" / "smart-greenhouse-bom.xlsx"
OUTPUT_CSV = REPO_ROOT / "docs" / "smart-greenhouse-bom.csv"

COLUMNS = [
    "Category",
    "Component",
    "Quantity",
    "Unit Price",
    "Previous Price",
    "Savings",
    "Store",
    "Currency",
    "Total Price",
    "Link / URL",
    "Notes",
]

# BOM: (category, component, qty, unit_price, currency, link, notes, prev_price, store)
# Prices re-checked across RU stores (Ozon, Chipdip, Youbot, Mean Well dealers, etc.), Jun 2026
BOM = [
    # 1.1 Серверная часть (Home Assistant)
    (
        "Серверная часть (Home Assistant)",
        "Raspberry Pi 5 4 GB",
        1,
        12900,
        "RUB",
        "https://www.iarduino.ru/shop/boards/raspberry-pi-5-4gb.html",
        "Рекомендуемый хост HA; док. 02 §1.1; оригинальная плата",
        14000,
        "iarduino.ru",
    ),
    (
        "Серверная часть (Home Assistant)",
        "Waveshare PoE M.2 HAT+ for Pi 5 (802.3at)",
        1,
        3290,
        "RUB",
        "https://onpad.ru/catalog/cubie/raspberrypi/boards/3770.html",
        "PoE + слот NVMe; док. 02 §1.1; оригинал Waveshare",
        4000,
        "onpad.ru",
    ),
    (
        "Серверная часть (Home Assistant)",
        "NVMe SSD 128 GB M.2 2230 (Kingston NV2 / WD SN580)",
        1,
        2100,
        "RUB",
        "https://www.ozon.ru/search/?text=NVMe+2230+128GB",
        "Загрузочный диск для HA OS; док. 02 §1.1; форм‑фактор 2230 редкий — проверить 2230 перед покупкой (не 2280)",
        2400,
        "Ozon",
    ),
    (
        "Серверная часть (Home Assistant)",
        "Raspberry Pi 5 Active Cooler + case",
        1,
        1700,
        "RUB",
        "https://tmelectronics.ru/product/SC1148/",
        "Охлаждение в сетевом шкафу; док. 02 §1.1; официальный Active Cooler (~841 ₽) + алюминиевый корпус (~859 ₽, onpad.ru)",
        2000,
        "tmelectronics.ru + onpad.ru",
    ),
    (
        "Серверная часть (Home Assistant)",
        "Patch cord Cat6 to network rack",
        0,
        0,
        "RUB",
        "https://www.chipdip.ru/product/l-ftp6-1-patch-kord-ftp-cat6-lszh-1m-linoya-9000934487",
        "Уже есть; Pi HA к PoE‑коммутатору; док. 02 §1.1 — кол‑во 0 к заказу",
        0,
        "",
    ),
    # 1.2 Сеть Mesh (уже есть — для полноты списка)
    (
        "Сеть Mesh",
        "Keenetic Speedster KN-3013 router",
        0,
        0,
        "RUB",
        "https://keenetic.com/ru/keenetic-speedster",
        "Уже есть; док. 02 §1.2 — кол‑во 0 к заказу",
        0,
        "",
    ),
    (
        "Сеть Mesh",
        "Netcraze Stellar 6 mesh AP (NAP-650 / KAP-650)",
        0,
        0,
        "RUB",
        "https://www.ozon.ru/search/?text=Netcraze+Stellar+6",
        "Уже есть; наружная PoE AP ~5 м от теплицы; док. 02 §1.2/01 §3",
        0,
        "",
    ),
    (
        "Сеть Mesh",
        "Keenetic PoE switch",
        0,
        0,
        "RUB",
        "https://keenetic.com/ru/catalog/poe-switches",
        "Уже есть; питает Pi HA + Stellar 6; док. 02 §1.2",
        0,
        "",
    ),
    (
        "Сеть Mesh",
        "Outdoor Cat6 cable to mesh AP",
        0,
        0,
        "RUB",
        "https://www.chipdip.ru/search?searchtext=кабель+Cat6+наружный",
        "Уже есть; ~15–20 м наружный Cat6 к mesh AP; док. 02 §1.2/01 §3 — кол‑во 0 к заказу",
        0,
        "",
    ),
    # 1.3 Щит IP65 снаружи у входа (не внутри теплицы)
    (
        "Щит IP65 у теплицы",
        "ESP32-DevKitC ESP32-WROOM-32U (U.FL)",
        2,
        595,
        "RUB",
        "https://www.youbot.ru/product/plata-esp32-devkit-wroom-32u",
        "greenhouse-watering + greenhouse-climate; док. 04 §1.3; разъём U.FL для внешней антенны",
        1000,
        "Youbot",
    ),
    (
        "Щит IP65 у теплицы",
        "Wi-Fi antenna 2.4 GHz 5 dBi + U.FL pigtail ~50 cm",
        2,
        180,
        "RUB",
        "https://alishops.ru/product-2-4-ggts-wifi-antenna-5dbi-antenna-rp-sma-shtekerniy-razem-2-4-ggts-antenna-wi-fi-marshrutizator-plus-21-sm-pci-u-fl-ipx-k-sma-shtirevoy-kabel/32972778284",
        "Внешние антенны вне металлического корпуса; док. 04 §1.3; реселлер AliExpress — доставка 2–4 нед. vs местный склад",
        450,
        "AliExpress (alishops.ru)",
    ),
    (
        "Щит IP65 у теплицы",
        "Mean Well LRS-50-5 (5 V 10 A)",
        1,
        812,
        "RUB",
        "https://www.meanwellrus.ru/catalog/ac_dc_bloki_pitaniya/lrs_seriya/lrs_50_5/",
        "Питание логики: ESP32, датчики, PCA9685 VCC; док. 04 §1.3/04 §2.5; оригинал Mean Well",
        2100,
        "meanwellrus.ru",
    ),
    (
        "Щит IP65 у теплицы",
        "Mean Well LRS-100-12 (12 V 8.5 A)",
        1,
        963,
        "RUB",
        "https://www.meanwellrus.ru/catalog/ac_dc_bloki_pitaniya/lrs_seriya/lrs_100_12/",
        "Питание реле и соленоидов; док. 04 §1.3/04 §2.5; оригинал Mean Well",
        2850,
        "meanwellrus.ru",
    ),
    (
        "Щит IP65 у теплицы",
        "Relay module 12 V opto-isolated 2-channel",
        2,
        210,
        "RUB",
        "https://compacttool.ru/modul-rele-2-kanalniy-12v-s-optorazvyazkoy",
        "По одному на ESP32 полива (клапаны полива и бака); док. 04 §1.3/03 §1.5; SRD-12VDC-SL-C + PC817",
        350,
        "CompactTool",
    ),
    (
        "Щит IP65 у теплицы",
        "PCA9685 16-channel PWM servo driver",
        1,
        365,
        "RUB",
        "https://amperkot.ru/spb/catalog/shim_kontroller_s_i2c_dlya_servoprivodov__plata_kontrollera_pwm_pca9685-24271675.html",
        "Драйвер MG996R на greenhouse-climate; док. 04 §1.3/04 §2.4",
        650,
        "Amperkot",
    ),
    (
        "Щит IP65 у теплицы",
        "Diode 1N4007 (flyback protection)",
        4,
        0,
        "RUB",
        "https://www.chipdip.ru/product/1n4007",
        "Уже в наличии по док.; по одному на катушку реле/соленоида; док. 04 §2.1",
        0,
        "",
    ),
    (
        "Щит IP65 у теплицы",
        "Enclosure DKC R5ST0342 IP65 300×400×200 mm (metal)",
        1,
        4458,
        "RUB",
        "https://electroprice.ru/product/shkaf-st-s-montazh--platoy-300h400h200mm-ot-ip65-do-ip66-ik10-dkc-r5st0342/117349",
        "Полевой щит снаружи у входа теплицы (не внутри RH 70–95%); док. 04 §1.3 / 03 §0; DKC R5ST0342 с монтажной платой",
        5500,
        "Electroprice",
    ),
    (
        "Щит IP65 у теплицы",
        "WAGO 221 terminals + cable glands PG7/PG9/PG11 + DIN rail kit",
        1,
        1800,
        "RUB",
        "https://www.chipdip.ru/search?searchtext=WAGO+221+набор",
        "1 комплект; WAGO 221 + сальники DKC PG + DIN‑рейка; док. 04 §1.3; оценка набора",
        2500,
        "Chipdip + IEK",
    ),
    (
        "Щит IP65 у теплицы",
        "FTP Cat5e outdoor cable (50 m roll)",
        1,
        2200,
        "RUB",
        "https://www.ozon.ru/search/?text=Extralink+Cat5e+FTP+outdoor+50m",
        "Из щита снаружи в теплицу (датчики, GPIO); doc. 04 §1.3/04 §2.5; Extralink EX.32341 или аналог DKC outdoor FTP",
        2800,
        "Ozon",
    ),
    (
        "Щит IP65 у теплицы",
        "Mini UPS 12 V DC 3 A (optional)",
        1,
        2400,
        "RUB",
        "https://www.ozon.ru/search/?text=mini+UPS+12V+DC+3A+WGP",
        "Опционально; удерживает клапаны в безопасном состоянии при отключении питания; док. 04 §1.3/01 §6.4; класс WGP V1203",
        3000,
        "Ozon",
    ),
    (
        "Щит IP65 у теплицы",
        "Gore vent IP67 (pressure equalization)",
        1,
        450,
        "RUB",
        "https://www.ozon.ru/search/?text=Gore+vent+IP67+электрощит",
        "Антиконденсат: щит снаружи у влажного входа; док. 04 §1.3/01 §5.1; вентиляционный винт IP67 (аналог Gore vent)",
        700,
        "Ozon",
    ),
    (
        "Щит IP65 у теплицы",
        "Circuit breaker 10 A (220 V AC input)",
        1,
        220,
        "RUB",
        "https://www.chipdip.ru/search?searchtext=автомат+10А+IEK",
        "Оценка; док. 04 §2.5/01 §6.4; IEK/DEK 1P 10 A",
        400,
        "Chipdip",
    ),
    (
        "Щит IP65 у теплицы",
        "PSU 5 V ≥3 A for servo V+ on PCA9685",
        1,
        550,
        "RUB",
        "https://www.chipdip.ru/search?searchtext=DL-15W-V5-EXC",
        "Отдельно от LRS-50-5; док. 04 §2.4/01 §5.2; закрытый БП DONE 5 V 3 A 15 W",
        1000,
        "Chipdip",
    ),
    (
        "Щит IP65 у теплицы",
        "Resistor 4.7 kΩ (I2C / OneWire pull-up)",
        0,
        0,
        "RUB",
        "https://www.chipdip.ru/search?searchtext=резистор+4.7+кОм",
        "Уже есть; подтяжки DS18B20 + I2C; док. 04 §2.1 — кол‑во 0 к заказу",
        0,
        "",
    ),
    (
        "Щит IP65 у теплицы",
        "Capacitor 100 nF (float switch debounce)",
        0,
        0,
        "RUB",
        "https://www.chipdip.ru/search?searchtext=конденсатор+100нФ",
        "Уже есть; опциональный debounce поплавка; док. 01 §5.2 — кол‑во 0 к заказу",
        0,
        "",
    ),
    # 1.4 Датчики
    (
        "Датчики",
        "SHT31 temperature/humidity module (GY-SHT31-D)",
        3,
        295,
        "RUB",
        "https://www.youbot.ru/product/tsifrovoy-datchik-temperatury-i-vlazhnosti-gy-sht31d",
        "Вход низ, центр наверху, противоположная сторона; I2C 0x44/0x45; док. 03 §1.4.1 / 04 §2.3",
        350,
        "Youbot",
    ),
    (
        "Датчики",
        "BH1750 light sensor module (GY-302)",
        2,
        119,
        "RUB",
        "https://www.youbot.ru/product/gy-302-bh1750-datchik-sveta",
        "Потолок + уровень растений; док. 03 §1.4",
        200,
        "Youbot",
    ),
    (
        "Датчики",
        "DS18B20 waterproof temperature probe (1 m)",
        2,
        190,
        "RUB",
        "https://www.yourduino.ru/product/vodonepronitsaemyy-temperaturnyy-datchik-zond-ds18b20",
        "Бак + линия полива; OneWire; док. 03 §1.4 / 04 §2.2",
        200,
        "YourDuino",
    ),
    (
        "Датчики",
        "Float switch level sensor (12–24 V NO/NC)",
        1,
        280,
        "RUB",
        "https://www.ozon.ru/search/?text=поплавковый+датчик+уровня+воды+12V",
        "Верхний уровень бака; док. 03 §1.4 / 04 §2.2; бытовой поплавок (не промышленный ПДУ)",
        350,
        "Ozon",
    ),
    (
        "Датчики",
        "YF-S201 water flow sensor 1/2 inch",
        2,
        248,
        "RUB",
        "https://www.youbot.ru/product/datchik-rashoda-vody-yf-s201",
        "Учёт полива и наполнения бака; док. 03 §1.4",
        450,
        "Youbot",
    ),
    # 1.5 Исполнительные устройства
    (
        "Исполнительные устройства",
        "Solenoid valve NC 12 V 1/2 inch brass",
        2,
        617,
        "RUB",
        "https://alishops.ru/product-ebowan-g1-2-latunniy-elektricheskiy-elektromagnitniy-klapan-n-c-12v-24v-220v-g3-4-vodniy-vozdushniy-vpusknoy-pereklyuchatel-potoka-dlya-solnechnogo-vodonagrevatelya-klapan/32917349764",
        "Полив + наполнение бака; fail-safe NC; док. 03 §1.5; латунь 12 V DC прямого действия; реселлер AliExpress — проверить NC и мин. давление 0 bar",
        900,
        "AliExpress (alishops.ru)",
    ),
    (
        "Исполнительные устройства",
        "MG996R servo (metal gears)",
        2,
        320,
        "RUB",
        "https://www.youbot.ru/product/servoprivod-mg996r-180-gradusov-metall",
        "Привод форточек через PCA9685; док. 03 §1.5 / 04 §2.4; MG996R 180° с металлическими шестернями",
        450,
        "Youbot",
    ),
    (
        "Монтаж и запас",
        "Drip tape 16 mm emitter spacing 20 cm roll 100 m",
        1,
        590,
        "RUB",
        "https://lemanapro.ru/catalog/search?text=капельная+лента+16+мм+20+см+100",
        "Капельный полив после YF-S201; док. 03 §2.6.4; Neo-Drip или аналог; + штуцер 1/2→16 мм и фильтр",
        1500,
        "Lemana PRO",
    ),
    # 1.7 Computer vision (optional — Phase 0, not in base ~52k)
    (
        "Computer vision (optional)",
        "PoE IP camera 5 MP IP66/IP67 (Reolink RLC-410 or analog)",
        2,
        7500,
        "RUB",
        "https://www.ozon.ru/search/?text=Reolink+RLC-410+PoE",
        "CV‑1 коридор у входа + CV‑2 излом П; RTSP; док. 05 §3 / 03 §1.4.2; не в базовом BOM",
        9000,
        "Ozon",
    ),
    (
        "Computer vision (optional)",
        "Outdoor Cat6 to greenhouse (cameras, if no spare)",
        1,
        2500,
        "RUB",
        "https://www.chipdip.ru/search?searchtext=кабель+Cat6+наружный",
        "30–40 m к PoE‑коммутатору; shared conduit с ESP32 или отдельная трасса; док. 05 §3.4",
        3200,
        "Chipdip",
    ),
    (
        "Computer vision (optional)",
        "GORE vent + drip loop kit for cameras (×2)",
        1,
        1000,
        "RUB",
        "https://www.ozon.ru/search/?text=Gore+vent+IP67+камера",
        "Антиконденсат на корпусе; doc. 05 §3.3",
        1600,
        "Ozon",
    ),
    (
        "Computer vision (optional)",
        "Radxa ZERO 3W 1GB RAM / 8GB eMMC (edge CV host)",
        1,
        3790,
        "RUB",
        "https://onpad.ru/catalog/cubie/radxa/rock/radxa_zero_series/3542.html",
        "greenhouse-cv-edge · .13 · RTSP + ONNX pre-filter · local post to Pi; no WAN; doc. 05 §3.4 / 04 §1.4",
        4200,
        "onpad.ru",
    ),
    (
        "Computer vision (optional)",
        "USB-C buck 5V 3A (edge SBC power from LRS-50-5)",
        1,
        450,
        "RUB",
        "https://www.chipdip.ru/search?searchtext=USB-C+5V+3A+buck",
        "Питание edge SBC от щита; doc. 04 §2.5 / 05 §3.4",
        600,
        "Chipdip",
    ),
    (
        "Computer vision (optional)",
        "microSD 32GB (edge SBC backup, optional with eMMC)",
        1,
        500,
        "RUB",
        "https://www.ozon.ru/search/?text=microSD+32GB+Class10",
        "Резервная загрузка Radxa; doc. 05 §3.4 — опционально при 8GB eMMC",
        700,
        "Ozon",
    ),
    # 1.6 Монтаж и запас
    (
        "Монтаж и запас",
        "Cable ties, ferrules, wire, fuses, silica gel, misc.",
        1,
        3000,
        "RUB",
        "https://www.chipdip.ru/search?searchtext=кабельные+стяжки+наконечник",
        "Док. 02 §1.6; incl. silica gel 50–100 g для щита снаружи (doc. 04 §1.3); ~10% запас",
        3500,
        "Chipdip + Ozon",
    ),
]


def line_savings(qty, unit_price, prev_price):
    if not qty or not prev_price or not unit_price:
        return ""
    return (prev_price - unit_price) * qty


def build_rows():
    rows = []
    for category, component, qty, unit_price, currency, link, notes, prev_price, store in BOM:
        total = qty * unit_price if qty else 0
        savings = line_savings(qty, unit_price, prev_price)
        rows.append(
            [
                category,
                component,
                qty,
                unit_price if unit_price else "",
                prev_price if prev_price else "",
                savings if savings != "" else "",
                store,
                currency,
                total if total else (0 if qty == 0 else ""),
                link,
                notes,
            ]
        )
    return rows


def write_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    header_fill = PatternFill("solid", fgColor="366092")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(COLUMNS)
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    purchasable_total = sum(
        r[8] for r in rows if isinstance(r[8], (int, float)) and r[8] > 0
    )
    prev_total = sum(
        (r[4] * r[2])
        for r in rows
        if isinstance(r[4], (int, float))
        and isinstance(r[2], (int, float))
        and r[2] > 0
        and r[4] > 0
    )
    total_savings = prev_total - purchasable_total if prev_total else 0

    ws.append([])
    ws.append(
        [
            "",
            "ИТОГО (к заказу, qty > 0)",
            "",
            "",
            prev_total,
            total_savings,
            "",
            "RUB",
            purchasable_total,
            "",
            "Без уже имеющегося сетевого оборудования (qty 0); цены проверены июнь 2026",
        ]
    )
    total_row = ws.max_row
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    ws.cell(row=total_row, column=9).font = Font(bold=True)

    widths = [22, 48, 10, 12, 14, 12, 22, 10, 12, 55, 45]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        link_cell = row[9]
        if link_cell.value and str(link_cell.value).startswith("http"):
            link_cell.hyperlink = str(link_cell.value)
            link_cell.font = Font(color="0563C1", underline="single")

    meta = wb.create_sheet("About")
    meta["A1"] = "Спецификация умной теплицы (BOM)"
    meta["A1"].font = Font(bold=True, size=14)
    meta["A3"] = "Источник"
    meta["B3"] = "docs/02-components-and-server.md, 03-greenhouse-installation.md, 04-esp32-and-cabinet.md (§1.1–1.6), 05-computer-vision.md (optional CV); index: smart-greenhouse-design.md"
    meta["A4"] = "Сгенерировано"
    meta["B4"] = "scripts/generate_bom.py"
    meta["A5"] = "Основа цен"
    meta["B5"] = "Сравнение цен магазинов, средний сегмент, июнь 2026 (ранее — сбалансированные ★ оценки)"
    meta["A6"] = "Валюта"
    meta["B6"] = "RUB, если не указано иное"
    meta["A7"] = "Предыдущий итог (к заказу)"
    meta["B7"] = f"{prev_total:,} RUB"
    meta["A8"] = "Текущий итог (к заказу)"
    meta["B8"] = f"{purchasable_total:,} RUB"
    meta["A9"] = "Общая экономия"
    meta["B9"] = f"{total_savings:,} RUB"
    meta["A11"] = "Условные обозначения"
    meta["B11"] = "Unit/Total Price = 0 при qty 0 — уже есть, не заказывать"
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 70

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


def write_csv(rows):
    import csv

    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        writer.writerows(rows)


def main():
    rows = build_rows()
    write_xlsx(rows)
    write_csv(rows)
    purchasable = sum(r[8] for r in rows if isinstance(r[8], (int, float)) and r[8] > 0)
    prev = sum(
        (r[4] * r[2])
        for r in rows
        if isinstance(r[4], (int, float))
        and isinstance(r[2], (int, float))
        and r[2] > 0
        and r[4] > 0
    )
    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_CSV}")
    print(f"Components: {len(rows)}")
    print(f"Previous purchasable total (RUB): {prev:,}")
    print(f"Current purchasable total (RUB): {purchasable:,}")
    print(f"Total savings (RUB): {prev - purchasable:,}")


if __name__ == "__main__":
    main()
